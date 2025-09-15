import time
from keypad_code import getData1  # Importing the keypad code
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime  # Import datetime module for timestamping
import RPi.GPIO as GPIO
from Adafruit_IO import Client, Feed
import subprocess

import smbus2
import time
from RPLCD.i2c import CharLCD

# Replace 'YOUR_AIO_USERNAME' and 'YOUR_AIO_KEY' with your actual Adafruit IO username and key
AIO_USERNAME = 'Aditya715'  # Your Adafruit IO username
AIO_KEY = 'aio_EraE02ZOZGDTr6rNqZvwp6uTYFLB'            # Your Adafruit IO Key

# Initialize the Adafruit IO client
aio = Client(AIO_USERNAME, AIO_KEY)

# The feed name you want to send data to (make sure this feed exists in your Adafruit IO account)
code = 'code'
m1 = 'm1'
m2 = 'm2'
m3 = 'm3'
m4 = 'm4'

# Set up the I2C LCD
lcd = CharLCD(i2c_expander='PCF8574', address=0x27, port=1, cols=16, rows=2)


# Define GPIO pin numbers for the switches
switches = {
    'a': 27,     # GPIO pin for switch 'a'
    's': 22,     # GPIO pin for switch 's'
    'd': 23,     # GPIO pin for switch 'd'
    'f': 24,     # GPIO pin for switch 'f'
    'submit': 25 # GPIO pin for submit switch
}

#motor pins
motor_1 = 8
motor_2 = 1
motor_3 = 0
motor_4 = 7


# Set up GPIO pins for motor control
GPIO.setup(motor_1, GPIO.OUT)
GPIO.setup(motor_2, GPIO.OUT)
GPIO.setup(motor_3, GPIO.OUT)
GPIO.setup(motor_4, GPIO.OUT)


# Set the motor to be enabled initially
GPIO.output(motor_1, GPIO.LOW)
GPIO.output(motor_2, GPIO.LOW)
GPIO.output(motor_3, GPIO.LOW)
GPIO.output(motor_4, GPIO.LOW)


'''
#motor test
GPIO.output(motor_1, GPIO.HIGH)
GPIO.output(motor_2, GPIO.HIGH)
GPIO.output(motor_3, GPIO.HIGH)
GPIO.output(motor_4, GPIO.HIGH)

time.sleep(2)
#motor test
GPIO.output(motor_1, GPIO.LOW)
GPIO.output(motor_2, GPIO.LOW)
GPIO.output(motor_3, GPIO.LOW)
GPIO.output(motor_4, GPIO.LOW)

'''

# Initialize GPIO
GPIO.setmode(GPIO.BCM)
GPIO.setup(list(switches.values()), GPIO.IN, pull_up_down=GPIO.PUD_UP)

lcd.clear()
lcd.cursor_pos = (0, 0)
lcd.write_string(f"Medicine")
lcd.cursor_pos = (1, 0)
lcd.write_string(f"Dispenser")
time.sleep(2)

# Create or load the Excel workbook
def create_or_load_workbook(filename):
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Add header for new workbook
        sheet.append(["S.No", "Date and Time", "USN", "a", "s", "d", "f"])
        workbook.save(filename)
    return workbook

# Save barcode data, date, time, and medicine counts to Excel
def save_to_excel(filename, barcode, medicine_counts, serial_number):
    workbook = create_or_load_workbook(filename)
    sheet = workbook.active
    
    # Get current date and time
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Add the barcode entry with the serial number and date-time
    sheet.append([serial_number, current_time, barcode,
                 medicine_counts['a'],
                 medicine_counts['s'],
                 medicine_counts['d'],
                 medicine_counts['f']])
    
    # Save the updated Excel file
    workbook.save(filename)
    print(f"Data saved for barcode {barcode}")

# Main function to capture input from scanner, switches, and save to Excel
def capture_barcode_and_medicine_data():
    filename = "medicine.xlsx"
    serial_number = 1  # Start serial number from 1 or read it from the existing Excel sheet
    
    # Load existing workbook to get the latest serial number
    workbook = create_or_load_workbook(filename)
    sheet = workbook.active
    if sheet.max_row > 1:
        serial_number = sheet.max_row  # Continue from the last used serial number
    
    while True:
        try:
            # Step 1: Scan the barcode
            print("Scan a barcode:")
            lcd.clear()
            lcd.cursor_pos = (0, 0)
            lcd.write_string(f"Scan a barcode:")
            barcode = input()  # Simulating barcode scanner input
            if barcode:
                print(f"Barcode scanned: {barcode}")
                lcd.clear()
                lcd.cursor_pos = (0, 0)
                lcd.write_string(f"Barcode scanned:")
                lcd.cursor_pos = (1, 0)
                lcd.write_string(f"{barcode}")
                time.sleep(2)
                # Step 2: Initialize medicine counts
                medicine_counts = {'a': 0, 's': 0, 'd': 0, 'f': 0}
                lcd.clear()
                lcd.cursor_pos = (0, 0)
                lcd.write_string(f"m1")
                lcd.cursor_pos = (0, 3)
                lcd.write_string(f"m2")
                lcd.cursor_pos = (0, 7)
                lcd.write_string(f"m3")
                lcd.cursor_pos = (0, 11)
                lcd.write_string(f"m4")
                
                # Step 3: Detect medicine switches
                print("Press switches for tablets. Press 'submit' to finish.")
                while True:
                    for medicine, pin in switches.items():
                        if GPIO.input(pin) == GPIO.LOW:  # Button press detected
                            if medicine == 'submit':
                                # On submit, save the data and increment the serial number
                                save_to_excel(filename, barcode, medicine_counts, serial_number)
                                serial_number += 1  # Increment serial number for the next entry
                                time.sleep(1)  # Debounce delay
                                lcd.clear()
                                lcd.cursor_pos = (0, 0)
                                lcd.write_string(f"data submitted")
                                lcd.cursor_pos = (1, 0)
                                lcd.write_string(f"Stored in excel")
                                time.sleep(2)
                                # Call the Adafruit upload 
                                print("data uploading...")
                                lcd.clear()
                                lcd.cursor_pos = (0, 0)
                                lcd.write_string(f"data uploading..")
                                aio.send(code, barcode)
                                print(f'Successfully sent {code} to feed {barcode}')
                                time.sleep(1)
                                aio.send(m1, medicine_counts['a'])
                                print(f'Successfully sent {medicine_counts["a"]} to feed {m1}')
                                time.sleep(1)
                                aio.send(m2, medicine_counts['s'])
                                print(f'Successfully sent {medicine_counts["s"]} to feed {m2}')
                                time.sleep(1)
                                aio.send(m3, medicine_counts['d'])
                                print(f'Successfully sent {medicine_counts["d"]} to feed {m3}')
                                time.sleep(1)
                                aio.send(m4, medicine_counts['f'])
                                print(f'Successfully sent {medicine_counts["f"]} to feed {m4}')
                                time.sleep(1)
                                print("Completed")
                                lcd.cursor_pos = (1, 0)
                                lcd.write_string(f"data uploaded")
                                time.sleep(2)
                                break
                            else:
                                # Increment count for the respective medicine
                                medicine_counts[medicine] += 1
                                print(f"{medicine} count: {medicine_counts[medicine]}")
                                lcd.cursor_pos = (1, 0)
                                lcd.write_string(f"{medicine_counts['a']}")
                                lcd.cursor_pos = (1, 3)
                                lcd.write_string(f"{medicine_counts['s']}")
                                lcd.cursor_pos = (1, 7)
                                lcd.write_string(f"{medicine_counts['d']}")
                                lcd.cursor_pos = (1, 11)
                                lcd.write_string(f"{medicine_counts['f']}")

                                time.sleep(1)  # Debounce delay
                                if medicine =='a':
                                    GPIO.output(motor_1, GPIO.HIGH)
                                    time.sleep(0.3)
                                    GPIO.output(motor_1, GPIO.LOW)
                                if medicine =='s':
                                    GPIO.output(motor_2, GPIO.HIGH)
                                    time.sleep(0.3)
                                    GPIO.output(motor_2, GPIO.LOW)
                                if medicine =='d':
                                    GPIO.output(motor_3, GPIO.HIGH)
                                    time.sleep(0.3)
                                    GPIO.output(motor_3, GPIO.LOW)
                                if medicine =='f':
                                    GPIO.output(motor_4, GPIO.HIGH)
                                    time.sleep(0.3)
                                    GPIO.output(motor_4, GPIO.LOW)

                                
                    else:
                        continue  # Continue inner loop if submit not pressed
                    break  # Exit the outer loop if submit is pressed
        
        except KeyboardInterrupt:
            print("\nProcess interrupted.")
            break

def main():
    print("Main program started")
    lcd.clear()
    lcd.cursor_pos = (0, 0)
    lcd.write_string(f"Enter password:")

    input_data = getData1()  # Waits for password input from keypad
    # Continuously prompt for input
    lcd.cursor_pos = (1, 0)
    lcd.write_string(f"{input_data}")
    time.sleep(1)
    while True:
        OTP = input_data
        print(f"Entered password: {OTP}") 
        # Example validation logic for the entered password
        if OTP == "715":  # Check if the entered password is correct
            lcd.clear()
            lcd.cursor_pos = (0, 0)
            lcd.write_string(f"Welcome")
            time.sleep(2)
            #print("Password correct!")
            capture_barcode_and_medicine_data()
            #break  # Exit the loop if password is correct
        else:
            print("Incorrect password, please try again.")
            lcd.clear()
            lcd.cursor_pos = (0, 0)
            lcd.write_string(f"Incorrect password:")
            lcd.cursor_pos = (1, 0)
            lcd.write_string(f"Enter again")
            time.sleep(2)
            lcd.clear()
            lcd.cursor_pos = (0, 0)
            lcd.write_string(f"Re-Enter password:")
            input_data = getData1()  # Waits for password input from keypad
            lcd.cursor_pos = (1, 3)
            lcd.write_string(f"{input_data}")  
            time.sleep(1)  # Optional: Delay to avoid rapid re-entering of input

    print("Main program continues")

# Example of running the main program
if __name__ == "__main__":
    main()

